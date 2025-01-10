# -*- coding=utf-8-*-
from Crypto.Cipher import AES
from binascii import b2a_hex, a2b_hex
import hashlib
from crcmod import *
from binascii import *
import binascii
import base64
import xlrd
import configparser
import logging
import os
import time
from xlutils.copy import copy
import openpyxl
import xlwt
from Crypto.Util.Padding import pad
from Crypto.Util.Padding import unpad
import pymysql
import re


# 例子如下
# 明文:str(binascii.b2a_hex(str.encode('tuya49368a48b746')))[2:-1]
# 密文:51506024af73925e15da8873afb08f9d
# key:str(binascii.b2a_hex(str.encode('8sHRRhqNAdXnSvpA')))[2:-1]
# iv:str(binascii.b2a_hex(str.encode('8sHRRhqNAdXnSvpA')))[2:-1]

# 如果text不足16位的倍数就用"00"补足,即采用NoPadding方式补齐
# def PKCS_zero(text):
#     newbytes = '00'
#     if len(text) % 32:
#         add = 32 - (len(text) % 32)
#         add = add >> 1
#     else:
#         add = 0
#     text = text + newbytes * add
#     return text

# 数据库读取_查询
def select(sql, num=-1):
    print(sql)
    host = get_config("db", "host")
    print(host)
    user = get_config("db", "user")
    password = get_config("db", "password")
    database = get_config("db", "database")
    port = get_config("db", "port")
    con = pymysql.connect(host=host, port=int(port), user=user, password=password, database=database, charset="utf8")
    logger.info("开始连接数据库--host:" + host + ",user:" + user + ",password:" + password)
    cur = con.cursor()
    cur.execute("use " + database)

    # 取出全部结果
    if num == -1:
        try:
            print(sql)
            cur.execute(sql)
            source = cur.fetchall()
            # logger.info("查询sql:%s" % sql + "，查询结果:%s" % source)
            con.commit()
            cur.close()
            con.close()
            # print(source)
            return source
        except TypeError as e:
            # except NameError as e:
            logger.error("查询出错：%s" % e)
            cur.close()
            con.close()
    # 取出指定数量
    else:
        try:
            cur.execute(sql)
            source = cur.fetchmany(num)
            logger.info("查询sql:%s" % sql + "，查询结果:%s" % source)
            con.commit()
            cur.close()
            con.close()
            return source
        except:
            logger.error("查询出错sql：%s" % sql)
            cur.close()
            con.close()


# 加密函数
def AES_CBC_encrypt(text, key, iv):
    bs = 16
    PADDING = lambda s: s + (bs - len(s) % bs) * chr(bs - len(s) % bs)
    mode = AES.MODE_CBC
    # text = PKCS_zero(text)
    # text = bytes.fromhex(text)
    # print("plain :", bytes.hex(text), type(text))
    key = bytes.fromhex(key)
    # print("key:%s"%key)
    iv = bytes.fromhex(iv)
    cryptos = AES.new(key, mode, iv)
    # print("1",cryptos)
    crypt = cryptos.encrypt(PADDING(text).encode('utf-8'))
    crypted_str = base64.b64encode(crypt)

    # # cipher_text = bytes.hex(cryptos.encrypt(text))
    # print(cipher_text)
    # 因为AES加密后的字符串不一定是ascii字符集的，输出保存可能存在问题，所以这里转为16进制字符串

    return crypted_str
    # 解密后，去掉补足的空格用strip() 去掉


# AES-CBC解密
def AES_CBC_decrypt(text, key, iv):
    # print("AES_CBC_decrypt")
    # print(" key  :", key, type(key))
    # print(" iv   :", iv, type(iv))
    # print("plain :", text, type(text))
    bs = 16
    # PADDING = lambda s: s + (bs - len(s) % bs) * chr(bs - len(s) % bs)
    # text = bytes.fromhex(text)
    key = bytes.fromhex(key)
    print(key)
    iv = bytes.fromhex(iv)
    print(iv)
    # padtext = pad(bytes.fromhex(text), 16, style='pkcs7')
    text = base64.b64decode(text)
    print(text)
    mode = AES.MODE_CBC

    cryptos = AES.new(key, mode, iv)

    # cryptos = AES.new(key, mode, iv)
    # crypt = cryptos.encrypt(PADDING(text).encode('utf-8'))
    # crypted_str = base64.b64encode(crypt)

    # plain_text = bytes.hex(cryptos.decrypt(text))
    plain_text = cryptos.decrypt(text)
    plain_text1 = plain_text.hex()
    # print("plain_text1:",plain_text1)
    # print("cipher:", plain_text, type(plain_text))
    # print("************************************************")
    return plain_text


def is_11_digit_number(s):
    pattern = r'^\d{15}$'
    return bool(re.match(pattern, s))


# 读取excel数据
class ExcelOperate():
    def __init__(self, filename, filename1):
        self.filename = filename
        self.filename1 = filename1

    # 读取sheet表
    def read_excel(self, sheetname):
        book = xlrd.open_workbook(filename=self.filename)
        sheet = book.sheet_by_name(sheetname)
        return sheet

    # 获取一行的数据
    def get_excel_data(self, sheetname, raw):
        sheet = self.read_excel(sheetname)
        row = sheet.row_values(raw)
        return row

    def get_nrows(self, sheetname):
        sheetname = self.read_excel(sheetname)
        return sheetname.nrows

    def get_sheetnames(self):
        book = xlrd.open_workbook(self.filename)
        return book.sheet_names()

    def getdict(self, str):
        mydict = {}
        data = str.split(",")
        for d in data:
            newdata = d.split(":")
            mydict[newdata[0]] = newdata[1]
        return mydict

    def copy_excel(self):
        workbook = xlrd.open_workbook(self.filename)
        new_workbook = copy(workbook)
        new_workbook.save(self.filename1)
        return new_workbook

    def write_excel(self, sheetname, row, col, value):
        workbook = xlrd.open_workbook(self.filename)
        # new_workbook = copy(workbook)
        new_worksheet = workbook.get_sheet(sheetname)
        # new_worksheet = workbook.sheet_by_name(sheetname)
        # sheet = self.read_excel(sheetname)
        new_worksheet.write(row, col, value)
        workbook.save(self.filename1)  # 保存文件
        workbook.close()

    def write_excel1(self, sheetname, row, col, value):
        wb = openpyxl.load_workbook(self.filename1)
        sh = wb[sheetname]
        wb.close()
        sh.cell(row=row, column=col, value=value)
        wb.save(self.filename1)

    def write(self, sheetname, i, j, value):
        # self.excel_create()
        # rb = xlrd.open_workbook(self.filename)
        # wb = copy(rb)  # 管道作用，通过get_sheet()获取的sheet有write()方法
        ws = self.copy_excel().get_sheet(sheetname)  # 1代表是写到第几个工作表里，从0开始算是第一个。
        ws.write(i, j, value)
        ws.save(self.filename)

    def write_value_open(self, sheetname, row, col, value):
        # 先用xlrd打开源文件

        wb = openpyxl.load_workbook(self.filename)
        work_sheet = wb[sheetname]
        # work_sheet['K2']='pass' #此种写法后面可能容易报错，保险起见我使用了下面的方法
        work_sheet.cell(row=row, column=col).value = value
        wb.save(self.filename)


# def write_excel_xls_append(self, value,sheetname):
#     """写入excel"""
#     index = len(value)  # 获取需要写入数据的行数
#     workbook = xlrd.open_workbook(filename=self.filename)  # 打开工作簿
#     # sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
#     worksheet = workbook.sheet_by_name(sheetname)  # 获取工作簿中所有表格中的的第一个表格
#     rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
#     new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
#     new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
#     for i in range(0, index):
#         for j in range(0, len(value[i])):
#             new_worksheet.write(i, j, value[i][j])  # 从第一行开始写
#             # new_worksheet.write(i + rows_old, j, value[i][j])  # 追加写入数据，注意是从i+rows_old行开始写入
#     new_workbook.save(path)  # 保存工作簿
#     print("xls格式表格【覆盖数据前%s行】写入数据成功！" % index)
#
# def get_importData(self, num):
#     """获取需写入excel的数据
#     @:param num 写入的条数
#     """
#     datalist = [["手机号", "车牌号"]]  # excel的第一行表头
#     for i in range(num):
#         phone = self.random.get_Phonenum()
#         car = self.random.get_CarPlatnumber()
#         value = ['%s' % phone, '%s' % car]
#         datalist.append(value)
#     return datalist


# 读取配置文件
def get_config(aa, bb):
    conf = configparser.ConfigParser()
    file_path = os.path.dirname(os.path.abspath('..')) + '/Electronic student/config.ini'
    conf.read(file_path, encoding='UTF-8')
    cc = conf.get(aa, bb)
    return cc


class Logger(object):
    ch = logging.StreamHandler()

    def __init__(self, logger):
        '''指定保存日志的文件路径，日志级别以及调用文件将日志存入到指定的文件中'''
        # 创建一个logger
        self.logger = logging.getLogger(logger)
        self.logger.setLevel(logging.DEBUG)
        # 创建一个handler，用于写入日志文件
        rq = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        # log_path = os.path.dirname(os.path.abspath('.')) + '\Electronic student\\'
        # log_path = os.path.dirname(os.path.abspath('.')) + '\\'
        log_path = os.path.abspath('..') + '\\'
        # 如果case组织结构式/testsuit/fraturemoddel/xxx.py，那么得到的相对路径的父路径就是项目的根目录
        log_name = log_path + rq + '.log'
        fh = logging.FileHandler(log_name)
        fh.setLevel(logging.INFO)

        # 再创建一个handler，用于输出到控制台
        self.ch.setLevel(logging.INFO)
        # 定义handler的输出格式
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        self.ch.setFormatter(formatter)
        # 给logger添加handler
        self.logger.addHandler(fh)
        self.logger.addHandler(self.ch)

    def debug(self, message):
        self.fontColor('\033[0;32m%s\033[0m')
        self.logger.debug(message)

    def info(self, message):
        self.fontColor('\033[0;34m%s\033[0m')
        self.logger.info(message)

    def warning(self, message):
        self.fontColor('\033[0;37m%s\033[0m')
        self.logger.warning(message)

    def error(self, message):
        self.fontColor('\033[0;31m%s\033[0m')
        self.logger.error(message)

    def critical(self, message):
        self.fontColor('\033[0;35m%s\033[0m')
        self.logger.critical(message)

    def fontColor(self, color):
        # 不同的日志输出不同的颜色
        formatter = logging.Formatter(color % '[%(asctime)s] - [%(levelname)s] - %(message)s')
        self.ch.setFormatter(formatter)
        self.logger.addHandler(self.ch)

    def getlog(self):
        return self.logger


logger = Logger(logger="until").getlog()


# 字符串分割
def str_split(str, key):
    aa = str.split("#")
    return aa[key]


# aa传入文本，bb第一次切割，cc第二次切割，dd第一次需要位置下标,
def str_split1(aa, bb, cc, dd, ee):
    text = aa.split(bb)
    text1 = text[dd].split(cc)
    return text1[ee]


def str_split2(aa, bb):
    aa1 = aa.split(bb)
    return aa1


class date_factory():
    pass

    def date_head(self):
        pass


class analysis_date():
    def DEVICE_LOGIN(self, text):
        a = str_split1(text, ",", "]", 6, 0)
        print(a)
        if a == "0@0@0":
            logger.info("登录成功，响应报文:%s" % str(a))
            return "登录成功，响应报文:%s" % str(a)
        if a == "1@0@0":
            logger.info("注册失败，响应报文:%s" % str(a))
            return "注册失败，响应报文:%s" % str(a)
        else:
            str1 = str_split2(a, "@")
            if str1[0] == "0":
                if str1[2] == "0":
                    logger.info("注册成功，端口号:%s，不发送短信" % str(str1[1]))
                    return "注册成功，端口号:%s，不发送短信" % str(str1[1])
                if str1[2] == "1":
                    logger.info("注册成功，端口号:%s，发送短信" % str(str1[1]))
                    return "注册成功，端口号:%s，发送短信" % str(str1[1])
            if str1[0] == "1":
                logger.info("非平台用户")
                return "非平台用户"
            if str1[0] == "2":
                logger.info("设置异常")
                return "设置异常"

    def currency(self, text):
        a = str_split1(text, ",", "]", 7, 0)
        if a == "0":
            return "正常"
        if a == "1":
            return "非平台用户"
        if a == "2":
            return "其他异常"
        else:
            return "错误"

    def judge(self, text):
        a = str_split2(text, ",")
        if a[3] == "DEVICE_LOGIN":
            self.DEVICE_LOGIN(text)
            return self.DEVICE_LOGIN(text)
        if a[
            3] == "ALARM_POWER" or "REPORT_CROSS_BORDER" or "REPORT_LOCATION_INFO" or "REPORT_HEARTBEAT" or "REPORT_CALL_LOG" or "REPORT_SOS" or "DEVICE_STATUS" or "REPORT_SMS_READ" or "REPORT_DEVICE_INFO":
            self.currency(text)
            return self.currency(text)
    # def ALARM_POWER(self,text):
    #     a = str_split1(text,",","]",6,0)
    #     if a == 0:
    #         logger.info("正常" )
    #     if a == 1:
    #         logger.info("非平台用户")
    #     if a == 2:
    #         logger.info("其他异常")
    # def REPORT_CROSS_BORDER(self,text):
    #     a = str_split1(text, ",", "]", 6, 0)
    #     if a == 0:
    #         logger.info("正常")
    #     if a == 1:
    #         logger.info("非平台用户")
    #     if a == 2:
    #         logger.info("其他异常")
    # def REPORT_LOCATION_INFO(self,text):
    #     a = str_split1(text, ",", "]", 6, 0)
    #     if a == 0:
    #         logger.info("正常")
    #     if a == 1:
    #         logger.info("非平台用户")
    #     if a == 2:
    #         logger.info("其他异常")
    # def REPORT_HEARTBEAT(self,text):
    #     a = str_split1(text, ",", "]", 6, 0)
    #     if a == 0:
    #         logger.info("正常")
    #     if a == 1:
    #         logger.info("非平台用户")
    #     if a == 2:
    #         logger.info("其他异常")
    # def REPORT_CALL_LOG(self,text):
    #     a = str_split1(text, ",", "]", 6, 0)
    #     if a == 0:
    #         logger.info("正常")
    #     if a == 1:
    #         logger.info("非平台用户")
    #     if a == 2:
    #         logger.info("其他异常")
    # def REPORT_SOS(self,text):
    #     a = str_split1(text, ",", "]", 6, 0)
    #     if a == 0:
    #         logger.info("正常")
    #     if a == 1:
    #         logger.info("非平台用户")
    #     if a == 2:
    #         logger.info("其他异常")
    # def DEVICE_STATUS(self,text):
    #     a = str_split1(text, ",", "]", 6, 0)
    #     if a == 0:
    #         logger.info("正常")
    #     if a == 1:
    #         logger.info("非平台用户")
    #     if a == 2:
    #         logger.info("其他异常")
    # def REPORT_SMS_READ(self,text):
    #     a = str_split1(text, ",", "]", 6, 0)
    #     if a == 0:
    #         logger.info("正常")
    #     if a == 1:
    #         logger.info("非平台用户")
    #     if a == 2:
    #         logger.info("其他异常")
    # def REPORT_DEVICE_INFO(self,text):
    #     a = str_split1(text, ",", "]", 6, 0)
    #     if a == 0:
    #         logger.info("正常")
    #     if a == 1:
    #         logger.info("非平台用户")
    #     if a == 2:
    #         logger.info("其他异常")


if __name__ == '__main__':
    key = "96 B6 71 5E F5 0F A4 55 7F 6C F9 77 17 8E 86 C9"
    iv = "11 C5 00 74 0B E4 4D 4E E5 BD AE D0 3C E7 6F FF"
    res = AES_CBC_encrypt(
        "[862677060102391,89860315243719513266,202306271026100000,DEVICE_LOGIN,4,20230627102610,5,0@0@0]", key, iv)
    # res1 = AES_CBC_encrypt("[890890890,89860484012080099999,202108281421200000,REPORT_LOCATION_INFO,3,20210828142120,79,0E114.008739N22.392291T20210828142120@460!0!9231!2351@wifi!AC:BC:32:78:A2:5F!-97]", key, iv)
    print("res加密:%s" % res, type(res))
    # print("res1加密:%s" % res1, type(res))
    res1 = AES_CBC_decrypt(
        "t9xNZ5l+otyRJ8kgWo6uGcrelEJFBvPdirtLi5DZRbc/GzkEnJCGcciSSavjRzQ0XiBo5CyFwgCSZah8LsgCLpHJTYxqQCSnMAmOTlSgbq8wKIyY0j9HZwetgHLDVgTHDUAJ0o+2/24cvF+exk8eABvUh4gcgH8UXnSWrkvPU0M=)",
        key, iv)
    print("res1:%s" % res1)
    data = b'\u5361\u53f7'
    print(data.decode())
    # sql = "select schoolId from " + "tbl_school WHERE orgId = 1"
    # print(sql)
    # aa = select("select classId from tbl_class WHERE schoolId = 499")
    # print(len(aa))
